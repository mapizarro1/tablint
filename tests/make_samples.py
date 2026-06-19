"""Generate messy sample files to exercise tablint checks."""

import csv
import os

import openpyxl
from openpyxl.utils import get_column_letter

OUT = os.path.join(os.path.dirname(__file__), "samples")
os.makedirs(OUT, exist_ok=True)


def w_csv(name, rows, delimiter=","):
    path = os.path.join(OUT, name)
    with open(path, "w", newline="", encoding="utf-8") as fh:
        wr = csv.writer(fh, delimiter=delimiter)
        wr.writerows(rows)
    return path


def main():
    # 1. Clean CSV.
    w_csv(
        "clean.csv",
        [["id", "name", "age"], [1, "Ann", 30], [2, "Bob", 41], [3, "Cy", 22]],
    )

    # 2. Header on row 4 (preamble rows above).
    w_csv(
        "header_row4.csv",
        [
            ["Quarterly Export"],
            ["Generated 2026-06-01"],
            [],
            ["id", "name", "amount"],
            [1, "Ann", "100"],
            [2, "Bob", "200"],
        ],
    )

    # 3. Numbers stored as text (quoted thousands).
    w_csv(
        "numbers_as_text.csv",
        [
            ["sku", "qty", "price"],
            ["A1", "1,000", "12.50"],
            ["A2", "2,500", "9.99"],
            ["A3", "3,200", "15.00"],
            ["A4", "900", "7.25"],
            ["A5", "1,100", "8.00"],
        ],
    )

    # 4. Ragged rows + a totals row.
    w_csv(
        "ragged_totals.csv",
        [
            ["region", "q1", "q2"],
            ["East", 10, 20],
            ["West", 15],  # ragged
            ["North", 12, 18, 99],  # ragged extra
            ["Total", 37, 38],
        ],
    )

    # 5. Duplicate and empty column names.
    w_csv(
        "dup_empty_cols.csv",
        [
            ["id", "name", "name", ""],
            [1, "Ann", "A", "x"],
            [2, "Bob", "B", "y"],
        ],
    )

    # 6. Ambiguous dates.
    w_csv(
        "ambiguous_dates.csv",
        [
            ["id", "date"],
            [1, "03/04/2026"],
            [2, "05/06/2026"],
            [3, "11/12/2026"],
        ],
    )

    # 7. TSV clean.
    w_csv(
        "clean.tsv",
        [["id", "city"], [1, "NYC"], [2, "LA"]],
        delimiter="\t",
    )

    # 8. Multiple tables stacked on one XLSX sheet + hidden sheet + merged + formula + error.
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    # Table A
    ws.append(["id", "name", "amount"])
    ws.append([1, "Ann", 100])
    ws.append([2, "Bob", 200])
    ws.append([])  # blank separator
    ws.append([])
    # Table B
    ws.append(["code", "desc"])
    ws.append(["X", "alpha"])
    ws.append(["Y", "beta"])
    # merged cells
    ws.merge_cells("A12:C12")
    ws["A12"] = "merged note"
    # formula and error
    ws["E1"] = "=SUM(C2:C3)"
    ws["E2"] = "=1/0"
    # hidden sheet
    hs = wb.create_sheet("Secret")
    hs["A1"] = "hidden data"
    hs.sheet_state = "hidden"
    wb.save(os.path.join(OUT, "multi_table.xlsx"))

    # 9. Clean XLSX single table.
    wb2 = openpyxl.Workbook()
    ws2 = wb2.active
    ws2.title = "Sheet1"
    ws2.append(["id", "name", "score"])
    for i in range(1, 11):
        ws2.append([i, f"user{i}", i * 3])
    wb2.save(os.path.join(OUT, "clean.xlsx"))

    # 10. Not a table: prose text file with .csv extension.
    with open(os.path.join(OUT, "prose.csv"), "w", encoding="utf-8") as fh:
        fh.write("This is just a paragraph of text. It has no delimiters at all.\n")
        fh.write("Another sentence here without any structure whatsoever.\n")

    print("samples written to", OUT)
    for f in sorted(os.listdir(OUT)):
        print(" ", f)


if __name__ == "__main__":
    main()
