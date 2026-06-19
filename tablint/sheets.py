"""XLSX loading and sheet-level extraction via openpyxl.

Bounded by row and sheet caps. Produces, per sheet, a row matrix plus the
xlsx-specific signals (hidden, merged cells, formula/error cells, external
links). Pure-ish: it reads a local file, no network, no payment awareness.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from typing import Any, List, Optional

import openpyxl


@dataclass
class SheetData:
    name: str
    hidden: bool
    matrix: List[List[Any]] = field(default_factory=list)
    n_rows: int = 0
    n_cols: int = 0
    merged_cells: int = 0
    formula_cells: int = 0
    error_cells: int = 0
    external_links: int = 0
    truncated: bool = False


def load_xlsx(
    file_path: str, max_sheets: int = 20, max_rows: int = 5000, max_cols: int = 200
) -> List[SheetData]:
    # Pass 1: structure + formulas + errors (data_only=False keeps formulas).
    wb_f = openpyxl.load_workbook(file_path, data_only=False, read_only=False)
    # Pass 2: cached values for analysis (data_only=True).
    wb_v = openpyxl.load_workbook(file_path, data_only=True, read_only=True)

    # External links are workbook level in openpyxl.
    ext_links = 0
    try:
        ext_links = len(getattr(wb_f, "_external_links", []) or [])
    except Exception:
        ext_links = 0

    out: List[SheetData] = []
    for sheet_name in wb_f.sheetnames[:max_sheets]:
        ws_f = wb_f[sheet_name]
        ws_v = wb_v[sheet_name] if sheet_name in wb_v.sheetnames else None

        hidden = getattr(ws_f, "sheet_state", "visible") != "visible"
        merged = len(getattr(ws_f, "merged_cells", []).ranges) if getattr(ws_f, "merged_cells", None) else 0

        formula_cells = 0
        error_cells = 0
        n_rows = 0
        n_cols = 0

        # Scan formulas/errors on the formula workbook (bounded).
        for r_i, row in enumerate(ws_f.iter_rows(max_row=max_rows, max_col=max_cols)):
            for cell in row:
                dt = getattr(cell, "data_type", None)
                if dt == "f":
                    formula_cells += 1
                elif dt == "e":
                    error_cells += 1

        # Build the value matrix from the data_only workbook (preferred), else
        # fall back to formula workbook values.
        matrix: List[List[Any]] = []
        src = ws_v if ws_v is not None else ws_f
        truncated = False
        for r_i, row in enumerate(src.iter_rows(values_only=True, max_row=max_rows + 1, max_col=max_cols)):
            if r_i >= max_rows:
                truncated = True
                break
            vals = list(row)
            matrix.append(vals)
            n_rows = max(n_rows, r_i + 1)
            n_cols = max(n_cols, len(vals))

        # Trim fully blank trailing rows.
        while matrix and all(v is None or (isinstance(v, str) and v.strip() == "") for v in matrix[-1]):
            matrix.pop()

        # Trim trailing all-blank columns. read_only iteration over-pads rows to
        # max_col with None; collapse to the true used width so structure and
        # column analysis are not polluted by phantom empty columns.
        true_width = 0
        for r in matrix:
            for ci in range(len(r) - 1, -1, -1):
                v = r[ci]
                if not (v is None or (isinstance(v, str) and v.strip() == "")):
                    true_width = max(true_width, ci + 1)
                    break
        matrix = [r[:true_width] for r in matrix]

        n_rows = len(matrix)
        n_cols = true_width

        out.append(
            SheetData(
                name=sheet_name,
                hidden=hidden,
                matrix=matrix,
                n_rows=n_rows,
                n_cols=n_cols,
                merged_cells=merged,
                formula_cells=formula_cells,
                error_cells=error_cells,
                external_links=ext_links,
                truncated=truncated,
            )
        )

    wb_v.close()
    wb_f.close()
    return out


def count_sheets(file_path: str) -> int:
    wb = openpyxl.load_workbook(file_path, read_only=True)
    try:
        return len(wb.sheetnames)
    finally:
        wb.close()
